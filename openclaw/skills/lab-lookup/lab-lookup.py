#!/usr/bin/env python3
"""Lab reference lookup — queries 03_Lab/network_info.xlsx for physical layer context.

Usage: python3 lab-lookup.py <command> <arg> [--xlsx <path>] [--lab-root <path>]

Commands:
  port-map <hostname>      Switch port, VLAN, patchpanel for a device
  nic-config <hostname>    NIC interfaces, bonds, VLANs, IPs
  vlan-devices <vlan_id>   All devices on a VLAN
  switch-ports <switch>    All populated ports on a switch
  docs <hostname>          List reference files in 03_Lab for a host
  ups-pdu <site>           UPS and PDU port assignments (nl or gr)
"""

import sys
import os
import glob

try:
    import openpyxl
except ImportError:
    print("No data found: openpyxl not installed")
    sys.exit(0)

DEFAULT_XLSX = "/app/reference-library/Cross-Site/network_info.xlsx"
DEFAULT_LAB_ROOT = "/app/reference-library"

# Sheet name patterns for switch port sheets
SWITCH_SHEETS = {
    "nl-sw01": ["NL01 - Cisco 3750X-48P-E", "NL01 - Cisco 3750X-48P-E v2"],
    "nlsw02": ["NL01 - Cisco 3750X-48P-E v2"],
    "gr-sw01": ["GR01 - Cisco CBS350-16T-E-2G"],
    "gr-sw02": ["GR01 - Cisco C1000-24T-4X-L"],
    "gr2sw01": ["GR02 - Cisco SG300-10 Ports"],
}

# Site prefix to sheet prefix mapping
SITE_PREFIXES = {
    "nllei": ["NL01 - ", "NL - "],
    "gr": ["GR01 - ", "GR - "],
    "gr2": ["GR02 - "],
}


def load_workbook(xlsx_path):
    if not os.path.exists(xlsx_path):
        return None
    return openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)


def detect_site(hostname):
    hostname_lower = hostname.lower()
    if hostname_lower.startswith("gr2"):
        return "gr2"
    if hostname_lower.startswith("grskg"):
        return "gr"
    return "nllei"


def get_site_sheet_prefixes(hostname):
    site = detect_site(hostname)
    return SITE_PREFIXES.get(site, ["NL01 - ", "NL - "])


def cmd_port_map(wb, hostname):
    """Search all switch and VLAN sheets for a hostname."""
    hostname_lower = hostname.lower()
    results = []

    for name in wb.sheetnames:
        ws = wb[name]
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            if not any(c is not None for c in row):
                continue
            cells = [str(c).strip() if c else "" for c in row]
            # Search for hostname in any cell
            match = False
            for cell in cells:
                if hostname_lower in cell.lower():
                    match = True
                    break
            if match:
                # Format based on sheet type
                if any(kw in name for kw in ["Cisco", "SG300", "Huawei", "ASA 5506"]):
                    # Switch port sheet
                    port = cells[0] if len(cells) > 0 else "?"
                    desc = cells[1] if len(cells) > 1 else ""
                    dev_port = cells[2] if len(cells) > 2 else ""
                    mode_idx = next((j for j, h in enumerate(headers) if "mode" in h.lower()), -1)
                    vlan_idx = next((j for j, h in enumerate(headers) if "vlan" in h.lower()), -1)
                    mode = cells[mode_idx] if mode_idx >= 0 and mode_idx < len(cells) else ""
                    vlan = cells[vlan_idx] if vlan_idx >= 0 and vlan_idx < len(cells) else ""
                    results.append(f"[{name}] Port: {port}, Device: {desc}, DevPort: {dev_port}, Mode: {mode}, VLAN: {vlan}")
                elif "vlan" in name.lower() or "mgmt" in name.lower() or "k8s" in name.lower():
                    # VLAN/IP sheet
                    ip_idx = next((j for j, h in enumerate(headers) if "ip" in h.lower()), 0)
                    host_idx = next((j for j, h in enumerate(headers) if "hostname" in h.lower()), 4)
                    conn_idx = next((j for j, h in enumerate(headers) if "connection" in h.lower()), 2)
                    port_idx = next((j for j, h in enumerate(headers) if "port" in h.lower() and "ip" not in h.lower()), -1)
                    ip = cells[ip_idx] if ip_idx < len(cells) else ""
                    host = cells[host_idx] if host_idx < len(cells) else ""
                    conn = cells[conn_idx] if conn_idx < len(cells) else ""
                    port = cells[port_idx] if port_idx >= 0 and port_idx < len(cells) else ""
                    results.append(f"[{name}] IP: {ip}, Host: {host}, Connection: {conn}, Port: {port}")
                elif "Patchpanel" in name:
                    port = cells[0] if len(cells) > 0 else "?"
                    device = cells[2] if len(cells) > 2 else cells[1] if len(cells) > 1 else ""
                    results.append(f"[{name}] Port: {port}, Device: {device}")
                elif "UPS" in name or "PDU" in name:
                    port = cells[0] if len(cells) > 0 else "?"
                    device = cells[1] if len(cells) > 1 else ""
                    results.append(f"[{name}] Port: {port}, Device: {device}")
                elif "Boot Order" in name:
                    # Skip boot order sheets for port-map
                    continue
                else:
                    # Generic row
                    results.append(f"[{name}] {' | '.join(c for c in cells[:6] if c)}")

    if not results:
        print(f"No data found for hostname '{hostname}' in network_info.xlsx")
    else:
        print(f"Physical layer data for {hostname} ({len(results)} matches):")
        for r in results:
            print(f"  {r}")


def cmd_nic_config(wb, hostname):
    """Look up the NIC sheet for a hostname."""
    sheet_name = f"{hostname} NIC"
    if sheet_name not in wb.sheetnames:
        # Try case-insensitive match
        for name in wb.sheetnames:
            if name.lower() == sheet_name.lower():
                sheet_name = name
                break
        else:
            print(f"No data found: no NIC sheet for '{hostname}' (available: {', '.join(n for n in wb.sheetnames if 'NIC' in n)})")
            return

    ws = wb[sheet_name]
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).strip() if c else "" for c in row]
        if i == 0:
            headers = cells
            continue
        if any(c for c in cells):
            rows.append(cells)

    if not rows:
        print(f"No data found: NIC sheet '{sheet_name}' is empty")
        return

    print(f"NIC config for {hostname} ({len(rows)} interfaces):")
    print(f"  {'Interface':<16} {'Type':<10} {'Parent':<10} {'Slaves/Ports':<20} {'IP Address':<18} {'VLAN':<6} {'Description'}")
    for cells in rows:
        iface = cells[0] if len(cells) > 0 else ""
        itype = cells[1] if len(cells) > 1 else ""
        parent = cells[2] if len(cells) > 2 else ""
        slaves = cells[3] if len(cells) > 3 else ""
        ip = cells[5] if len(cells) > 5 else ""
        # VLAN ID column varies; look for it in headers
        vlan_idx = next((j for j, h in enumerate(headers) if "vlan" in h.lower()), -1)
        vlan = cells[vlan_idx] if vlan_idx >= 0 and vlan_idx < len(cells) else ""
        desc_idx = next((j for j, h in enumerate(headers) if "description" in h.lower()), -1)
        desc = cells[desc_idx] if desc_idx >= 0 and desc_idx < len(cells) else ""
        if iface:
            print(f"  {iface:<16} {itype:<10} {parent:<10} {slaves:<20} {ip:<18} {vlan:<6} {desc}")


def cmd_vlan_devices(wb, vlan_id):
    """Find the VLAN sheet and list all devices."""
    vlan_id_str = str(vlan_id)
    matching_sheets = []

    for name in wb.sheetnames:
        name_lower = name.lower()
        if f"vlan{vlan_id_str})" in name_lower or f"vlan{vlan_id_str} " in name_lower or f"(vlan{vlan_id_str})" in name_lower:
            matching_sheets.append(name)

    if not matching_sheets:
        print(f"No data found: no VLAN sheet matching VLAN {vlan_id}")
        return

    for sheet_name in matching_sheets:
        ws = wb[sheet_name]
        headers = []
        devices = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c).strip() if c else "" for c in row]
            if i == 0:
                headers = cells
                continue
            if not any(c for c in cells):
                continue
            ip_idx = next((j for j, h in enumerate(headers) if "ip" in h.lower()), 0)
            host_idx = next((j for j, h in enumerate(headers) if "hostname" in h.lower()), 4)
            dev_idx = next((j for j, h in enumerate(headers) if "device" in h.lower()), 3)
            conn_idx = next((j for j, h in enumerate(headers) if "connection" in h.lower()), 2)
            addr_type_idx = next((j for j, h in enumerate(headers) if "type" in h.lower()), 1)
            ip = cells[ip_idx] if ip_idx < len(cells) else ""
            host = cells[host_idx] if host_idx < len(cells) else ""
            device = cells[dev_idx] if dev_idx < len(cells) else ""
            conn = cells[conn_idx] if conn_idx < len(cells) else ""
            addr_type = cells[addr_type_idx] if addr_type_idx < len(cells) else ""
            if ip and ip.lower() not in ("none", "ip address", ""):
                # Only include entries with a hostname or device name
                if host or device:
                    devices.append((ip, host, device, conn, addr_type))

        populated = [d for d in devices if d[1] or d[2]]
        print(f"[{sheet_name}] ({len(populated)} populated of {len(devices) + sum(1 for d in devices if not d[1] and not d[2])} total):")
        for ip, host, device, conn, addr_type in populated:
            host_str = f" ({host})" if host else ""
            print(f"  {ip:<18} {addr_type:<12} {conn:<10} {device}{host_str}")


def cmd_switch_ports(wb, switch_hostname):
    """List all ports on a switch."""
    switch_lower = switch_hostname.lower()

    # Direct hostname to sheet mapping
    sheet_names = SWITCH_SHEETS.get(switch_lower, [])

    # Fallback: search sheet names for the hostname
    if not sheet_names:
        for name in wb.sheetnames:
            if switch_lower in name.lower() or any(
                kw in name.lower() for kw in ["cisco", "sg300", "huawei"]
            ):
                if switch_lower[:8] in name.lower() or detect_site(switch_hostname) in name.lower()[:4]:
                    sheet_names.append(name)

    if not sheet_names:
        available = [n for n in wb.sheetnames if any(kw in n for kw in ["Cisco", "SG300", "Huawei"])]
        print(f"No data found: no switch sheet for '{switch_hostname}' (available: {', '.join(available)})")
        return

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = []
        ports = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c).strip() if c else "" for c in row]
            if i == 0:
                headers = cells
                continue
            if not any(c for c in cells):
                continue
            port = cells[0] if len(cells) > 0 else ""
            desc = cells[1] if len(cells) > 1 else ""
            dev_port = cells[2] if len(cells) > 2 else ""
            mode_idx = next((j for j, h in enumerate(headers) if "mode" in h.lower()), -1)
            vlan_idx = next((j for j, h in enumerate(headers) if "vlan" in h.lower()), -1)
            mode = cells[mode_idx] if mode_idx >= 0 and mode_idx < len(cells) else ""
            vlan = cells[vlan_idx] if vlan_idx >= 0 and vlan_idx < len(cells) else ""
            if port:
                ports.append((port, desc, dev_port, mode, vlan))

        print(f"[{sheet_name}] ({len(ports)} ports):")
        print(f"  {'Port':<28} {'Device':<30} {'DevPort':<16} {'Mode':<8} {'VLAN'}")
        for port, desc, dev_port, mode, vlan in ports:
            if desc:  # Only show populated ports
                print(f"  {port:<28} {desc:<30} {dev_port:<16} {mode:<8} {vlan}")


def cmd_docs(lab_root, hostname):
    """List reference files in 03_Lab for a hostname."""
    hostname_lower = hostname.lower()
    site = detect_site(hostname)

    search_paths = []
    if site == "nllei":
        search_paths = [
            os.path.join(lab_root, "NL", "Servers"),
            os.path.join(lab_root, "NL", "Inventory"),
            os.path.join(lab_root, "NL", "Changes"),
        ]
    elif site == "gr":
        search_paths = [
            os.path.join(lab_root, "GR", "gr", "Servers"),
            os.path.join(lab_root, "GR", "gr", "Inventory"),
            os.path.join(lab_root, "GR", "gr", "Changes"),
        ]
    elif site == "gr2":
        search_paths = [
            os.path.join(lab_root, "GR", "gr2", "Servers"),
            os.path.join(lab_root, "GR", "gr2", "Inventory"),
        ]

    # Also check cross-site
    search_paths.append(os.path.join(lab_root, "Cross-Site", "Servers"))

    found_files = []
    for base_path in search_paths:
        if not os.path.isdir(base_path):
            continue
        # Search for directories matching the hostname
        for entry in os.listdir(base_path):
            entry_path = os.path.join(base_path, entry)
            if hostname_lower in entry.lower() and os.path.isdir(entry_path):
                for root, dirs, files in os.walk(entry_path):
                    for f in files:
                        fpath = os.path.join(root, f)
                        rel = os.path.relpath(fpath, lab_root)
                        size = os.path.getsize(fpath)
                        if size > 1048576:
                            size_str = f"{size / 1048576:.1f}MB"
                        elif size > 1024:
                            size_str = f"{size / 1024:.0f}KB"
                        else:
                            size_str = f"{size}B"
                        found_files.append((rel, size_str))

    if not found_files:
        print(f"No data found: no reference docs for '{hostname}' in 03_Lab")
    else:
        print(f"Reference docs for {hostname} ({len(found_files)} files):")
        for rel, size in found_files[:30]:  # Cap at 30 files
            print(f"  {rel} ({size})")
        if len(found_files) > 30:
            print(f"  ... and {len(found_files) - 30} more files")


def cmd_ups_pdu(wb, site_arg):
    """List UPS and PDU port assignments for a site."""
    site_lower = site_arg.lower()

    if site_lower in ("nl", "nl", "nllei"):
        prefixes = ["NL01 - "]
    elif site_lower in ("gr", "gr", "grskg"):
        prefixes = ["GR01 - "]
    else:
        prefixes = ["NL01 - ", "GR01 - "]

    found = False
    for name in wb.sheetnames:
        if not any(name.startswith(p) for p in prefixes):
            continue
        if "UPS" not in name and "PDU" not in name and "Schucko" not in name:
            continue

        ws = wb[name]
        headers = []
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c).strip() if c else "" for c in row]
            if i == 0:
                headers = cells
                continue
            if any(c for c in cells):
                rows.append(cells)

        if rows:
            found = True
            print(f"[{name}] ({len(rows)} ports):")
            for cells in rows:
                port = cells[0] if len(cells) > 0 else ""
                device = cells[1] if len(cells) > 1 else ""
                if port:
                    extra = ""
                    if len(cells) > 2 and cells[2]:
                        extra = f" ({cells[2]})"
                    print(f"  {port}: {device}{extra}")

    if not found:
        print(f"No data found: no UPS/PDU sheets for site '{site_arg}'")


def main():
    args = sys.argv[1:]

    xlsx_path = DEFAULT_XLSX
    lab_root = DEFAULT_LAB_ROOT

    # Parse optional flags
    filtered = []
    i = 0
    while i < len(args):
        if args[i] == "--xlsx" and i + 1 < len(args):
            xlsx_path = args[i + 1]
            i += 2
        elif args[i] == "--lab-root" and i + 1 < len(args):
            lab_root = args[i + 1]
            i += 2
        else:
            filtered.append(args[i])
            i += 1
    args = filtered

    if len(args) < 2:
        print(__doc__)
        sys.exit(1)

    command = args[0]
    arg = args[1]

    if command == "docs":
        cmd_docs(lab_root, arg)
        return

    wb = load_workbook(xlsx_path)
    if wb is None:
        print(f"No data found: xlsx file not found at {xlsx_path}")
        return

    try:
        if command == "port-map":
            cmd_port_map(wb, arg)
        elif command == "nic-config":
            cmd_nic_config(wb, arg)
        elif command == "vlan-devices":
            cmd_vlan_devices(wb, arg)
        elif command == "switch-ports":
            cmd_switch_ports(wb, arg)
        elif command == "ups-pdu":
            cmd_ups_pdu(wb, arg)
        else:
            print(f"Unknown command: {command}")
            print(__doc__)
            sys.exit(1)
    finally:
        wb.close()


if __name__ == "__main__":
    main()
